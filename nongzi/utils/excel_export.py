"""Excel ????"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def export_to_excel(headers: list, rows: list, filename: str, sheet_name: str = "Sheet1") -> str:
    """?? .xlsx ?????????"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ????
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ???
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ???
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # ????
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell in row:
                if cell:
                    max_length = max(max_length, len(str(cell)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 40)

    # ????????
    export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)
    return filepath

def export_stock_count(stock_count, db_session) -> str:
    """导出盘点表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "盘点表"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"盘点单: {stock_count.order_no}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:F2")
    ws["A2"] = f"盘点人: {stock_count.counter_name}  日期: {stock_count.count_date.strftime('%Y-%m-%d') if stock_count.count_date else ''}"
    ws["A2"].font = Font(size=10)

    headers = ["商品编码", "商品名称", "规格", "单位", "系统库存", "实盘数"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, item in enumerate(stock_count.items, 5):
        vals = [item.product.code, item.product.generic_name, item.product.spec or "",
                item.product.base_unit, item.system_quantity, ""]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row:
                if cell_val:
                    max_length = max(max_length, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_length + 4, 40)

    export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
    os.makedirs(export_dir, exist_ok=True)
    filename = f"stock_count_{stock_count.order_no}.xlsx"
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)
    return filepath
